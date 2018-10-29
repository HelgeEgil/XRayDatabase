# -* b- coding: utf-8 -*-
"""
Created on Wed Aug 14 14:44:44 2013

@author: Testulf
"""

import openpyxl, numpy, os, datetime
#
#from Tkinter import Tk
#from tkFileDialog import askopenfilename
#Tk().withdraw()

from sql_connection import *

# connect to database
#db = MySQLdb.connect(host="127.0.0.1",
#                   user="helge",
#                   passwd="123bergen",
#                   db="mydb")

def xls2sql(files):
    cur = db.cursor()
    cur.execute('pragma foreign_keys=ON')
    
    for filename in files:
        filename = filename.decode('utf8') # don't crash if utf-8 in filename
        W = openpyxl.load_workbook(filename, use_iterators = True)
        p = W.get_sheet_by_name(name = "SQL")
        
        a = []
        
        number_of_rows = 2
        
        for row in p.iter_rows():
            for k in row:
                _v = k.internal_value
                a.append(_v)
        
        #qa_tags = [p.cell('A' + str(x)).value for x in range(1, 26)]
        #qa_values = [p.cell('B' + str(x)).value for x in range(1,26)]
        qa_dict = {}
        #for k, v in qa_tags, qa_values:
        #    qa_dict[k] = v
        
        number_of_columns = len(a) / 2
        aa = numpy.resize(a, [number_of_columns, number_of_rows])
        
        for row in aa:
            _a = unicode(row[0])
            _b = unicode(row[1])
            if ";" in _a:# or "%" in _a:
                print "INJECT ERROR, INPUT %s FOUND" % _a
                raise ImportError
            if ";" in _b:# or "%" in _b:
                print "INJECT ERROR, INPUT %s FOUND" % _b
                raise ImportError
            qa_dict[_a] = _b
        
        for k,v in qa_dict.items():
            print "k: {}, v: {}".format(k,v)
        
        people = qa_dict["ppl_names"].split(", ")
        doc_number = qa_dict["doc_number"]

        if not doc_number or doc_number == "None":
            # create one!
            # Format: YY-### (14-853 eg.)
            cur.execute("SELECT doc_number FROM qa ORDER BY doc_number DESC LIMIT 1")
            last_doc_number = cur.fetchall()[0]
            this_year = datetime.date.today().year
            this_year_format = str(this_year)[2:]
            
            if last_doc_number:
                last_year, last_num = last_doc_number.split("-")
                next_num = str(int(last_num) + 1)
                if not last_year == str(this_year):
                    print u"nytt år!"
                doc_number = "{}-{}".format(this_year_format, next_num)
            else:
                print "No previous document numbers found!"
                doc_number = "{}-001".format(this_year_format)
            print "Generated new document number: {}.\nPlease enter this into your Excel file... :)".format(doc_number)

                
        qa_date = qa_dict["qa_date"]
        serial = qa_dict["serial"].replace(".0", "")
        sys_id = qa_dict["system_id"]
        
        print "QA for machine", serial, "with filename", filename

        # find all people with abbreviated names in SQL        
        cur.execute("SELECT ppl_abbreviation, ppl_name FROM people WHERE ppl_abbreviation <> ''")
        replacelist = cur.fetchall()
        
        for replace_each in replacelist:
            people = [p.replace(replace_each[0], replace_each[1]) for p in people]
            
        print "People found in QA %s:\n -" % doc_number, "\n - ".join(people)
        
        # find SQL ID
        cur.execute("SELECT qa_id, FK_qa_mod FROM qa WHERE doc_number = ?", (doc_number,))
        result = cur.fetchall()
        if len(result) > 0:
            qa_id = result[0][0]
            mod_id = result[0][1]
        else:
            cur.execute("SELECT mod_id from modality WHERE serial = ? OR station_name = ?", (serial, sys_id))
            result = cur.fetchall()
            try:        
                mod_id = result[0][0]
            except IndexError:
                print "Could not find modality by serial number (%s) or station name (%s) in database..." % (serial, sys_id)
                break
            #print mod_id
            cur.execute("INSERT INTO qa(study_date, doc_number, FK_qa_mod) VALUES(?, ?, ?)", (qa_date, doc_number, mod_id))
            cur.execute("SELECT qa_id, FK_qa_mod FROM qa WHERE doc_number = ?", (doc_number,))
            result = cur.fetchall()
            qa_id = result[0][0]
            mod_id = result[0][1]
        
        cur.execute("SELECT study_date FROM qa WHERE FK_qa_mod = ? ORDER BY study_date DESC", (mod_id,))
        result = cur.fetchall()
        print
        print "Current QA date: %s" % result[0]
        if len(result) > 1:
            print "Last QA for machine: %s" % result[1]
        else:
            print "No former QA for machine found!"
            
        cur.execute("SELECT qa_data FROM qa WHERE qa_id = ?", (qa_id,))
        qa_data = cur.fetchall()[0][0]
        if qa_data:
            print "Data found for current QA, aborting.\n---------------------------\n"
            continue
        
        
    # Data management from file
    # Make dictionaries where keys = left hand column values in excel, and values = right hand column values
    # loop over every key, value pair in excel (they should be unique), and test against every dictionary key
    # if a match occurs, set the excel value (= right hand column) to the dictionary value.
    
    # In this format, it's easy to transfer the data to SQL (even with inequal SQL names)
    
        # FLUORO IMAGE QUALITY
        fiq_list = [
                "fiqe_lab_live_contrast",
                "fiqe_lab_lih_contrast",
                "fiqe_lab_live_lpmm",
                "fiqe_lab_lih_lpmm",
                "fiqe_control_live_contrast",
                "fiqe_control_lih_contrast",
                "fiqe_control_live_lpmm",
                "fiqe_control_lih_lpmm",
                "fiqe_pulsespeed",
                "fiqe_dosemode",
                "fiqe_fieldsize_cm",
                "FK_fiqe_fiq"
                ]
            
        fiq = {}
        for x in fiq_list:
            fiq[x] = [None]*6 # i - 1
        
        
        # DAP         
        dap_list = ["dap_each_hvl",
                    "dap_each_kv",
                    "dap_each_meter_mean_hvl",
                    "dap_each_meter_mean_kv",
                    "dap_each_modality_mean",
                    "FK_de_d"]
        
        dap_each = {}
        for x in dap_list:
            dap_each[x] = [None]*5 # i-1        
        
        # TUBE QA
        tqal_list = ["tqal_set_kvp",
                     "tqal_set_ma",
                     "tqal_set_ms",
                     "tqal_set_mas",
                     "tqal_meas_kvp",
                     "tqal_meas_microgy",
                     "tqal_meas_ms",
                     "tqal_meas_hvl_mmal",
                     "FK_tqal_tqa"]
        
        tqal = {}
        for x in tqal_list:
            tqal[x] = [None]*25
        
        tqas_list = ["tqas_set_kvp",
                     "tqas_set_ma",
                     "tqas_set_ms",
                     "tqas_set_mas",
                     "tqas_meas_kvp",
                     "tqas_meas_microgy",
                     "tqas_meas_ms",
                     "tqas_meas_hvl_mmal",
                     "FK_tqas_tqa"]
        
        tqas = {}
        for x in tqas_list:
            tqas[x] = [None]*25
        
        
        # FLUORO TUBE QA
        
        # FLUORO TUBE QA PARAMETERS
        ft_list = ["ftkp_comment",
                   "ftpxd_comment",
                   "ftdxd_comment",
                   "ftkp_program",
                   "ftkp_dosemode",
                   "ftkp_rasterfactor",
                   "ftkp_fka_cm",
                   "ftkp_fda_cm",
                   "ftpxd_internalfiltration",
                   "ftpxd_fha_cm",
                   "ftpxd_fka_cm",
                   "ftpxd_fda_cm",
                   "ftdxd_internalfiltration",
                   "ftdxd_rasterfactor",
                   "ftdxd_fka_cm",
                   "ftdxd_fda_cm"]
                   
        ft = {}
        for x in ft_list:
            ft[x] = None
        
        # fluoro tube kv precision large focus
        ftkpl_list = ["ftkpl_set_kv",
                      "ftkpl_meas_kv",
                      "ftkpl_meas_hvl_mmal",
                      "FK_ftkpl_ft"]
                      
        ftkpl = {}
        for x in ftkpl_list:
            ftkpl[x] = [None]*13
        
        # fluoro tube kv precision small focus
        ftkps_list = ["ftkps_set_kv",
                      "ftkps_meas_kv",
                      "ftkps_meas_hvl_mmal",
                      "FK_ftkps_ft"]
                      
        ftkps = {}
        for x in ftkps_list:
            ftkps[x] = [None]*5
        
        # fluoro tube patient dose measurement
        ftpxd_list = ["ftpxd_program",
                      "ftpxd_pps",
                      "ftpxd_dosemode",
                      "ftpxd_fieldsize_cm",
                      "ftpxd_panel_kv",
                      "ftpxd_panel_ma",
                      "ftpxd_panel_ms",
                      "ftpxd_meas_mgy_s",
                      "ftpxd_meas_pps",
                      "FK_ftpxd_ft"]
        
        ftpxd = {}
        for x in ftpxd_list:
            ftpxd[x] = [None]*10
        
        # fluoro tube detector dose measurement
        ftdxd_list = ["ftdxd_program",
                      "ftdxd_pps",
                      "ftdxd_dosemode",
                      "ftdxd_fieldsize_cm",
                      "ftdxd_panel_kv",
                      "ftdxd_panel_ma",
                      "ftdxd_panel_ms",
                      "ftdxd_meas_mgy_s",
                      "ftdxd_corrected_mgy_s",
                      "FK_ftdxd_ft"]
                      
        ftdxd = {}
        for x in ftdxd_list:
            ftdxd[x] = [None]*13
    
        # AEK AEK AEK AEK AEK AEK AEK AEK AEK
        aekt_list = ["aekt_raster",
                     "aekt_fda_cm",
                     "aekt_rasterfocus_cm",
                     "aekt_program",
                     "aekt_comment",
                     "aekt_kv",
                     "aekt_ma",
                     "aekt_focus",
                     "aekt_size",
                     "aekt_density",
                     "aekt_cal_slope",
                     "aekt_cal_offset",
                     "aekt_cov_1mm",
                     "aekt_cov_2mm"]
                                      
        aekt = {}
        for x in aekt_list:
            aekt[x] = None
    
        aekw_list = ["aekw_raster",
                     "aekw_fda_cm",
                     "aekw_rasterfocus_cm",
                     "aekw_program",
                     "aekw_kv",
                     "aekw_comment",
                     "aekw_ma",
                     "aekw_focus",
                     "aekw_size",
                     "aekw_density",
                     "aekw_cal_slope",
                     "aekw_cal_offset",
                     "aekw_cov_1mm",
                     "aekw_cov_2mm"]
        
        aekw = {}
        for x in aekw_list:
            aekw[x] = None
    
        aekt_cal_list = ["aektce_ms", "aektce_microgy"]
        aekt_cal = {}
        for x in aekt_cal_list:
            aekt_cal[x] = [None]*7
            
        aekw_cal_list = ["aekwce_ms", "aekwce_microgy"]
        aekw_cal = {}
        for x in aekw_cal_list:
            aekw_cal[x] = [None]*7
    
        aek_cal_list = ["aekc_slope_1mm", "aekc_offset_1mm",
                        "aekc_slope_2mm", "aekc_offset_2mm"]
    
        aek_cal = {}
        for x in aek_cal_list:
            aek_cal[x] = None
    
        aek_cal_each_list = ["aekce_ms_1mm", "aekce_microgy_1mm",
                             "aekce_ms_2mm", "aekce_microgy_2mm"]
    
        aek_cal_each = {}
        for x in aek_cal_each_list:
            aek_cal_each[x] = [None] * 7
    
    
        aekwe_list = ["aekwe_ms_1mm_ch1",
                      "aekwe_ms_1mm_ch2",
                      "aekwe_ms_1mm_ch3",
                      "aekwe_ms_2mm_ch1",
                      "aekwe_ms_2mm_ch2",
                      "aekwe_ms_2mm_ch3",
                      "aekwe_microgy_1mm_ch1",
                      "aekwe_microgy_1mm_ch2",
                      "aekwe_microgy_1mm_ch3",
                      "aekwe_microgy_2mm_ch1",
                      "aekwe_microgy_2mm_ch2",
                      "aekwe_microgy_2mm_ch3"]
                      
        aekwe = {}
        for x in aekwe_list:
            aekwe[x] = [None]*2
    
        aekte_list = ["aekte_ms_1mm_ch1",
                      "aekte_ms_1mm_ch2",
                      "aekte_ms_1mm_ch3",
                      "aekte_ms_2mm_ch1",
                      "aekte_ms_2mm_ch2",
                      "aekte_ms_2mm_ch3",
                      "aekte_microgy_1mm_ch1",
                      "aekte_microgy_1mm_ch2",
                      "aekte_microgy_1mm_ch3",
                      "aekte_microgy_2mm_ch1",
                      "aekte_microgy_2mm_ch2",
                      "aekte_microgy_2mm_ch3"]
                      
        aekte = {}
        for x in aekte_list:
            aekte[x] = [None]*2
    
        fiq_control_monitor = None
        fiq_lab_monitor = None
        tqa_comment = None
        qa_comment = None
        fiq_comment = None
        dap_comment = None
        ftkp_comment = None
        ftdxd_comment = None
        ftpxd_comment = None
        aekt_comment = None
        aekw_comment = None
        
        # LOOP THROUGH EXCEL FILE
        for k, v in qa_dict.items():
            # legacy support.....
            if "fiq_pulsespeed" in k or "fiq_dosemode" in k or "fiq_fieldsize" in k:
                k = k.replace("fiq_","fiqe_")
            if "fiqe_fieldsize" in k:
                k = k.replace("fiqe_fieldsize", "fiqe_fieldsize_cm")
                
            if "#VALUE" in v:
                v = None
            # FLUORO IMAGE QUALITY
            if "qa_comment" in k:
                if v and not v == u"0.0":
    #                print "Comment:", v
                    qa_comment = unicode(v)
        
            if "fiqe_comment" in k or "fiq_comment" in k:
                if v and not v == u"0.0":
                    fiq_comment = unicode(v)
                    
            if "dap_comment" in k:
                if v and not v == u"0.0":
                    dap_comment = unicode(v)
                    
            if "fiqe_control_monitor" in k:
                if v == u"ok":        
                    #print "Control monitor OK"
                    fiq_control_monitor = True
                elif v == u"0.0":
                    pass
                else:
                    #print "Control monitor not OK"
                    fiq_control_monitor = False
                    
            if "fiqe_lab_monitor" in k:
                if v == u"ok":        
                    #print "Lab monitor OK"
                    fiq_lab_monitor = True
                elif v == u"0.0":
                    pass
                else:
                    #print "Lab monitor not OK"
                    fiq_lab_monitor = False
                    
            for qk in fiq.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        fiq[qk][i-1] = None
                    else:
                        fiq[qk][i-1] = unicode(v)
        
            # DAP
            for qk in dap_each.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v or "#VALUE" in v:
                        dap_each[qk][i-1] = None
                    else:
                        dap_each[qk][i-1] = unicode(v)
        
            # TUBE QA
            
            if "tqa_fka_cm" in k:
                if v == u"0.0" or not v:
                    tqa_fka_cm = None
                else:
                    tqa_fka_cm = unicode(v)
                    
            if "tqa_filter_mmcu" in k:
                if v == u"0.0" or not v:
                    tqa_filter_mmcu = None
                else:
                    tqa_filter_mmcu = unicode(v)
                    
            if "tqa_comment" in k:
                if v == u"0.0" or not v:
                    tqa_comment = None
                else:
                    tqa_comment = unicode(v)#v.decode("utf-8")
            
            for qk in tqal.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        tqal[qk][i-1] = None
                    else:
                        tqal[qk][i-1] = unicode(v)
                        
            for qk in tqas.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        tqas[qk][i-1] = None
                    else:
                        tqas[qk][i-1] = unicode(v)
                        
            for qk in ft.keys():
                if qk in k:
                    if v == u"0.0" or not v:
                        ft[qk] = None
                    else:
                        ft[qk] = unicode(v)
                    
            for qk in ftkps.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        ftkps[qk][i-1] = None
                    else:
                        ftkps[qk][i-1] = unicode(v)
                        
            for qk in ftkpl.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        ftkpl[qk][i-1] = None
                    else:
                        ftkpl[qk][i-1] = unicode(v)
                        
            for qk in ftpxd.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        ftpxd[qk][i-1] = None
                    else:
                        ftpxd[qk][i-1] = unicode(v)
                        
            for qk in ftdxd.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        ftdxd[qk][i-1] = None
                    else:
                        ftdxd[qk][i-1] = unicode(v)
                        
            for qk in aekt.keys():
                if qk in k:
                    if v == u"0.0" or not v:
                        aekt[qk] = None
                    elif "slope" in k:
                            aek_cal["aekc_slope_1mm"] = unicode(v)
                    elif "offset" in k:
                            aek_cal["aekc_offset_1mm"] = unicode(v)                
                    else:
                        aekt[qk] = unicode(v)
        
            for qk in aekw.keys():
                if qk in k:
                    if v == u"0.0" or not v:
                        aekw[qk] = None
                    elif "slope" in k:
                        aek_cal["aekc_slope_2mm"] = unicode(v)
                    elif "offset" in k:
                        aek_cal["aekc_offset_2mm"] = unicode(v)
                    else:
                        aekw[qk] = unicode(v)
    
            # For legacy reasons - first editions used TABLE / WALL for calibration
            # instead of 1 mm / 2mm (which is definitely correct)
            
            for qk in aekt_cal.keys(): # 1 mm
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        aekt_cal[qk] = None
                    elif "ms" in k:
                        aek_cal_each["aekce_ms_1mm"][i-1] = unicode(v)
                    elif "microgy" in k:
                        aek_cal_each["aekce_microgy_1mm"][i-1] = unicode(v)
                        
            for qk in aekw_cal.keys(): # 2 mm
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        aekw_cal[qk][i-1] = None
                    elif "ms" in k:
                        aek_cal_each["aekce_ms_2mm"][i-1] = unicode(v)
                    elif "microgy" in k:
                        aek_cal_each["aekce_microgy_2mm"][i-1] = unicode(v)
    
            for qk in aek_cal.keys():
                if qk in k:
                    if v == u"0.0" or not v:
                        aek_cal[qk] = None
                    else:
                        aek_cal[qk] = unicode(v)
    
            for qk in aek_cal_each.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        aek_cal_each[qk][i-1] = None
                    else:
                        aek_cal_each[qk][i-1] = unicode(v)
                        
            for qk in aekte.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        aekte[qk][i-1] = None
                    else:
                        aekte[qk][i-1] = unicode(v)
                        
            for qk in aekwe.keys():
                if qk in k:
                    i = int(k.split("_")[-1])
                    if v == u"0.0" or not v:
                        aekwe[qk][i-1] = None
                    else:
                        aekwe[qk][i-1] = unicode(v)
                        
                        
                        
                        
                        
        ####################################
        # INSERT INTO DATABASE             #
        ####################################
    
        cur.execute("UPDATE qa SET qa_comment = ? WHERE qa_id = ?", (qa_comment, qa_id))
        cur.execute("UPDATE qa SET qa_data = 1 WHERE qa_id = ?", (qa_id,))
        
        for person in people:
            cur.execute("SELECT ppl_id FROM people WHERE ppl_name = ?", (person,))
            result = cur.fetchall()
            if len(result)>0:
                ppl_id = result[0][0]
            else:
                cur.execute("INSERT INTO people (ppl_name) VALUES (?)", (person,))
#                cur.execute("SELECT LAST_INSERT_ID()")
                ppl_id = cur.lastrowid # fetchall()[0][0]
            cur.execute("INSERT INTO qapeople (FK_qappl_ppl, FK_qappl_qa) VALUES (?, ?)", (ppl_id, qa_id))
        
        # FLUORO IMAGE QUALITY
        ####################################
        
        is_fiq = False
        for i in range(6): # if any of the 6 rows contains information, set is_fiq to True
            if sum(bool(fiq[fiq_list[j]][i]) for j in range(8)) > 0:
                # j in range(8) is where the data is
                is_fiq = True
        
        if is_fiq:
            cur.execute("SELECT count(FK_fiq_qa) FROM fluoro_iq WHERE FK_fiq_qa = ?", (qa_id,))
            n_fiq = cur.fetchall()[0][0]
            #if n_fiq>0: continue
            cur.execute("INSERT INTO fluoro_iq (fiq_lab_monitor, fiq_control_monitor, fiq_comment, FK_fiq_qa) VALUES (?, ?, ?, ?)", \
                    (fiq_lab_monitor, fiq_control_monitor, fiq_comment, qa_id))
#            cur.execute("SELECT LAST_INSERT_ID()")
            fiq_id = cur.lastrowid
            fiq["FK_fiqe_fiq"] = [fiq_id]*6
            
            for i in range(6):
                
                fiq_string = ", ".join(fiq_list)
                if sum(bool(fiq[fiq_list[j]][i]) for j in range(8)) == 0:
        #            print "Row", i+1, "of fluoro image quality is empty"
                    continue
                cur.execute("INSERT INTO fluoro_iq_each({}) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)".format(fiq_string),
                            [fiq[fiq_list[j]][i] for j in range(12)])
            
        # DAP
        ####################################
        
        isdap = False
        for i in range(5):
            if dap_each["dap_each_modality_mean"][i]:
                isdap = True
        
    #    if not isdap:
    #        print "No DAP found"
        
        if isdap:
            cur.execute("INSERT INTO dap (FK_dap_qa, dap_comment) VALUES (?, ?)", (qa_id, dap_comment))
#            cur.execute("SELECT LAST_INSERT_ID()")
            dap_id = cur.lastrowid # cur.fetchall()[0][0]
    
            dap_each["FK_de_d"] = [dap_id]*5
            
            for i in range(5):
                dap_string = ", ".join(dap_list)
                if (not dap_each["dap_each_modality_mean"][i]):
        #            print "Row", i+1, "of DAP is empty"
                    continue
                cur.execute("INSERT INTO dap_each({}) VALUES (?,?,?,?,?,?)".format(dap_string), 
                            [dap_each[dap_list[j]][i] for j in range(6)])
    
                
        # TUBE QA
        ####################################
        
        cur.execute("INSERT INTO tube_qa (FK_tqa_qa, tqa_fka_cm, tqa_filter_mmcu, tqa_comment) VALUES (?, ?, ?, ?)", 
                            (qa_id, tqa_fka_cm, tqa_filter_mmcu, tqa_comment))
#        cur.execute("SELECT LAST_INSERT_ID()")
        tqa_id = cur.lastrowid # fetchall()[0][0]
        
        # TUBE QA LARGE FOCUS
        ####################################
        
        tqal["FK_tqal_tqa"] = [tqa_id]*25
        
        for i in range(25):
            if sum(bool(tqal[tqal_list[j]][i]) for j in [4,5,6,7]) == 0:
    #            print "Row", i+1, "of TQAL is empty"
                continue
            tqal_string = ", ".join(tqal_list)
            cur.execute("INSERT INTO tube_qa_large ({}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)".format(tqal_string),
                        [tqal[tqal_list[j]][i] for j in range(9)])
                        
        # TUBE QA SMALL FOCUS
        ####################################
        
        tqas["FK_tqas_tqa"] = [tqa_id]*25
        
        for i in range(25):
            if sum(bool(tqas[tqas_list[j]][i]) for j in [4,5,6,7]) == 0:
    #            print "Row", i+1, "of TQAS is empty"
                continue
            tqas_string = ", ".join(tqas_list)
            cur.execute("INSERT INTO tube_qa_small ({}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)".format(tqas_string),
                        [tqas[tqas_list[j]][i] for j in range(9)])
                        
        # FLUORO TUBE QA
        ####################################
        
        ft_list.append("FK_ft_qa")
        ft["FK_ft_qa"] = qa_id
        
        if sum(bool(ft[ft_list[j]]) for j in range(len(ft_list))) == 0:
            continue
        
        ft_string = ", ".join(ft_list)
        cur.execute("INSERT INTO fluoro_tube ({}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, \
                                                            ?, ?, ?, ?, ?)".format(ft_string),
                    [ft[ft_list[j]] for j in range(len(ft_list))])
#        cur.execute("SELECT LAST_INSERT_ID()")
        ft_id = cur.lastrowid # fetchall()[0][0]
        
        
        # FLUORO TUBE kV PRECISION LARGE FOCUS
        ########################################################################
        
        ftkpl["FK_ftkpl_ft"] = [ft_id] * 13
        
        for i in range(13):
            if sum(bool(ftkpl[ftkpl_list[j]][i]) for j in [1,2]) == 0: # test on measured kvp + HVL
    #            print "Row", i+1, "of FTKPL is empty"
                continue
            ftkpl_string = ", ".join(ftkpl_list)
            cur.execute("INSERT INTO fluoro_tube_kvprec_large ({}) VALUES (?, ?, ?, ?)".format(ftkpl_string),
                        [ftkpl[ftkpl_list[j]][i] for j in range (4)])
       
        
        # FLUORO TUBE kV PRECISION SMALL FOCUS
        ########################################################################
    
        ftkps["FK_ftkps_ft"] = [ft_id] * 5
        
        ftkps_string = ", ".join(ftkps_list)
        for i in range(5):
            if sum(bool(ftkps[ftkps_list[j]][i]) for j in [1,2]) == 0: # test on measured kvp + HVL
    #            print "Row", i+1, "of FTKPS is empty"
                continue
            cur.execute("INSERT INTO fluoro_tube_kvprec_small ({}) VALUES (?, ?, ?, ?)".format(ftkps_string),
                        [ftkps[ftkps_list[j]][i] for j in range (4)])
        
        # FLUORO TUBE PATIENT SKIN DOSE
        ########################################################################
    
        ftpxd["FK_ftpxd_ft"] = [ft_id] * 10 # list since we loop over each ftpxd item below
        
        ftpxd_string = ", ".join(ftpxd_list)
        for i in range(10): # TEN measurements for each item
            if sum(bool(ftpxd[ftpxd_list[j]][i]) for j in [7,]) == 0: # measured mGy/s
    #            print "Row", i+1, "of FTPXD is empty"
                continue
            cur.execute("INSERT INTO fluoro_tube_pxdose ({}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)".format(ftpxd_string),
                        [ftpxd[ftpxd_list[j]][i] for j in range(10)])
        
        # FLUORO TUBE DETECTOR DOSE 
        ########################################################################
    
        ftdxd["FK_ftdxd_ft"] = [ft_id] * 13    
        
        ftdxd_string = ", ".join(ftdxd_list)
        for i in range(13):
            if sum(bool(ftdxd[ftdxd_list[j]][i]) for j in [8,]) == 0:
    #            print "Row", i+1, "of FTDXD is empty"
                continue
            cur.execute("INSERT INTO fluoro_tube_dxdose ({}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)".format(ftdxd_string),
                        [ftdxd[ftdxd_list[j]][i] for j in range(10)])
        # AEK     
        ####################################
         
        aekt_list = [_str.replace("aekt_", "aek_") for _str in aekt_list]
        aekw_list = [_str.replace("aekw_", "aek_") for _str in aekw_list]
        
        aekw_list.append("aek_position")
        aekt_list.append("aek_position")    
        
        # Exist for legacy reasons
        aekt_list.remove("aek_cal_slope")
        aekt_list.remove("aek_cal_offset")
        aekw_list.remove("aek_cal_slope")
        aekw_list.remove("aek_cal_offset")        
        
        for k, v in aekt.items():
            aekt[k.replace("aekt_", "aek_")] = v
        
        for k, v in aekw.items():
            aekw[k.replace("aekw_", "aek_")] = v
        
        aekt["aek_position"] = "table"
        aekw["aek_position"] = "wall"
        
    
        # AEK TABLE
        ####################################
        
        if sum(bool(aekt[aekt_list[j]]) for j in range(len(aekt_list))) == 1:
            print "No table AEK found"
        else:
            aekt_list.append("FK_aek_qa")
            aekt["FK_aek_qa"] = qa_id
            aekt_string = ", ".join(aekt_list)
            
            # minus to?
            cur.execute("INSERT INTO aek ({}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)".format(aekt_string),
                        [aekt[aekt_list[j]] for j in range(len(aekt_list))])
                        
#            cur.execute("SELECT LAST_INSERT_ID()")
            aekt_id = cur.lastrowid # fetchall()[0][0]
    
            microgy_1 = False
            microgy_2 = False
            ms_1 = False
            ms_2 = False
            for mmcu in ["1mm", "2mm"]:
                for chnum in ["ch1", "ch2", "ch3"]:
                    for k, v in aekte.items():
                        if mmcu in k and chnum in k:
                            if "microgy" in k:                        
                                microgy_1 = v[0]
                                microgy_2 = v[1]
                            if "ms" in k:
                                ms_1 = v[0]
                                ms_2 = v[0]
                            
                    cur.execute("INSERT INTO aek_each (aeke_mmcu, aeke_chamber_num, aeke_panel_ms_1, aeke_panel_ms_2, \
                                                       aeke_calc_dose_1, aeke_calc_dose_2, FK_aeke_aek) \
                                                       VALUES (?, ?, ?, ?, ?, ?, ?)",
                                                        (mmcu[0], chnum[-1], ms_1, ms_2, microgy_1, microgy_2, aekt_id))
    
        # AEK WALL
        ####################################
        
        if sum(bool(aekw[aekw_list[j]]) for j in range(len(aekw_list))) == 1:
            print "No wall AEK found"
        else:
            aekw_list.append("FK_aek_qa")
            aekw["FK_aek_qa"] = qa_id
            aekw_string = ", ".join(aekw_list)
            
            cur.execute("INSERT INTO aek ({}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)".format(aekw_string),
                        [aekw[aekw_list[j]] for j in range(len(aekw_list))])
            #cur.execute("SELECT LAST_INSERT_ID()")
            aekw_id = cur.lastrowid # fetchall()[0][0]
    
            microgy_1 = False
            microgy_2 = False
            ms_1 = False
            ms_2 = False
            for mmcu in ["1mm", "2mm"]:
                for chnum in ["ch1", "ch2", "ch3"]:
                    for k, v in aekwe.items():
                        if mmcu in k and chnum in k:
                            if "microgy" in k:
                                microgy_1 = v[0]
                                microgy_2 = v[1]
                            if "ms" in k:
                                ms_1 = v[0]
                                ms_2 = v[0]
                            
                    cur.execute("INSERT INTO aek_each (aeke_mmcu, aeke_chamber_num, aeke_panel_ms_1, aeke_panel_ms_2, \
                                                       aeke_calc_dose_1, aeke_calc_dose_2, FK_aeke_aek) \
                                                       VALUES (?, ?, ?, ?, ?, ?, ?)",
                                                        (mmcu[0], chnum[-1], ms_1, ms_2, microgy_1, microgy_2, aekw_id))
            
            # AEK DOSE METER CALIBRATION (mAs to microgy)
            ####################################
            
            
            cur.execute("INSERT INTO aek_calibration (aekc_mmcu, aekc_slope, aekc_offset, FK_aekc_qa) \
                VALUES (?, ?, ?, ?)",  ("1", aek_cal["aekc_slope_1mm"], aek_cal["aekc_offset_1mm"], qa_id))
                
#            cur.execute("SELECT LAST_INSERT_ID()")
            aekc_1mm_id = cur.lastrowid  #fetchall()[0][0]
                    
            cur.execute("INSERT INTO aek_calibration (aekc_mmcu, aekc_slope, aekc_offset, FK_aekc_qa) \
                VALUES (?, ?, ?, ?)", ("2", aek_cal["aekc_slope_2mm"], aek_cal["aekc_offset_2mm"], qa_id))    
#            cur.execute("SELECT LAST_INSERT_ID()")
            aekc_2mm_id = cur.lastrowid # fetchall()[0][0]
            
            
    #        aek_cal_each = aekce_ms_1mm, aekce_microgy_1mm, aekce_ms_2mm, aekce_microgy_2mm
            for i in range(7):
                if sum(bool(aek_cal_each[aek_cal_each_list[j]][i]) for j in range(len(aek_cal_each_list))) == 0:
    #                print "Row", i+1, "of table AEK is empty"
                    continue
                
                cur.execute("INSERT INTO aek_calibration_each (aekce_ms, aekce_microgy, FK_aekce_aekc) VALUES (?, ?, ?)",
                            (aek_cal_each["aekce_ms_1mm"][i], aek_cal_each["aekce_microgy_1mm"][i], aekc_1mm_id))
                            
                cur.execute("INSERT INTO aek_calibration_each (aekce_ms, aekce_microgy, FK_aekce_aekc) VALUES (?, ?, ?)",
                            (aek_cal_each["aekce_ms_2mm"][i], aek_cal_each["aekce_microgy_2mm"][i], aekc_2mm_id))
                            
    print "------------------------------------\n\nFinished reading all files."                        
    db.commit()
